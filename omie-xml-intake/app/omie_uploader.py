import os
import time
from datetime import datetime

import requests
from dotenv import load_dotenv
from openpyxl import load_workbook

from app.logging_utils import get_logger, log_section, log_status

OMIE_CLIENTES_URL = "https://app.omie.com.br/api/v1/geral/clientes/"
OMIE_CONTARECEBER_URL = "https://app.omie.com.br/api/v1/financas/contareceber/"
MIN_INTERVAL_SECONDS = 0.30


class RateLimiter:
    def __init__(self, min_interval_seconds: float):
        self.min_interval = min_interval_seconds
        self._last_ts = 0.0

    def wait(self) -> None:
        now = time.time()
        elapsed = now - self._last_ts
        if elapsed < self.min_interval:
            time.sleep(self.min_interval - elapsed)
        self._last_ts = time.time()


def _env_key(name: str, prefix: str) -> str:
    return f"{prefix}_{name}" if prefix else name


def _load_env() -> None:
    if os.path.exists(".env"):
        load_dotenv(dotenv_path=".env")
    else:
        load_dotenv()


def _safe_float(value) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).replace(",", ".").strip())
    except ValueError:
        return 0.0


def _iso_to_br(date_str: str) -> str:
    if not date_str:
        return ""
    try:
        return datetime.fromisoformat(date_str[:10]).strftime("%d/%m/%Y")
    except ValueError:
        return date_str


def _get_row_value(row, header_map, key: str) -> str:
    idx = header_map.get(key)
    if idx is None:
        return ""
    value = row[idx]
    return "" if value is None else str(value).strip()


def _build_request_body(call: str, payload: dict, prefix: str) -> dict:
    return {
        "call": call,
        "app_key": os.getenv(_env_key("OMIE_APP_KEY", prefix)),
        "app_secret": os.getenv(_env_key("OMIE_APP_SECRET", prefix)),
        "param": [payload],
    }


def _omie_request(url: str, call: str, payload: dict, rate_limiter: RateLimiter, prefix: str = "") -> dict:
    logger = get_logger("omie")
    rate_limiter.wait()
    body = _build_request_body(call, payload, prefix)
    if os.getenv(_env_key("OMIE_DRY_RUN", prefix), os.getenv("OMIE_DRY_RUN", "false")).lower() == "true":
        logger.info("[OMIE][DRY-RUN] %s %s", call, payload)
        return {"codigo_status": "0", "mensagem": "DRY RUN", "call": call}
    response = requests.post(url, json=body, timeout=30)
    response.raise_for_status()
    return response.json()


def _build_cliente_payload(row, header_map):
    payload = {
        "cnpj_cpf": _get_row_value(row, header_map, "destinatario_cnpj"),
        "razao_social": _get_row_value(row, header_map, "destinatario_nome"),
        "nome_fantasia": _get_row_value(row, header_map, "destinatario_nome"),
        "endereco": _get_row_value(row, header_map, "destinatario_xLgr"),
        "endereco_numero": _get_row_value(row, header_map, "destinatario_nro"),
        "complemento": _get_row_value(row, header_map, "destinatario_xCpl"),
        "bairro": _get_row_value(row, header_map, "destinatario_xBairro"),
        "cidade": _get_row_value(row, header_map, "destinatario_cMun") or _get_row_value(row, header_map, "destinatario_xMun"),
        "estado": _get_row_value(row, header_map, "destinatario_UF"),
        "cep": _get_row_value(row, header_map, "destinatario_CEP"),
    }
    return {k: v for k, v in payload.items() if v}


def _upsert_cliente(payload: dict, rate_limiter: RateLimiter, prefix: str):
    return _omie_request(OMIE_CLIENTES_URL, "UpsertClienteCpfCnpj", payload, rate_limiter, prefix)


def _incluir_conta_receber(row, header_map, codigo_cliente, rate_limiter: RateLimiter, prefix: str):
    codigo_categoria = os.getenv(_env_key("OMIE_CODIGO_CATEGORIA", prefix))
    id_conta_corrente = os.getenv(_env_key("OMIE_ID_CONTA_CORRENTE", prefix))
    if not codigo_categoria or not id_conta_corrente:
        raise ValueError("Configure OMIE_CODIGO_CATEGORIA e OMIE_ID_CONTA_CORRENTE no ambiente.")
    payload = {
        "codigo_lancamento_integracao": _get_row_value(row, header_map, "nfcom_id") or _get_row_value(row, header_map, "numero_nf"),
        "codigo_cliente_fornecedor": int(codigo_cliente) if codigo_cliente else None,
        "codigo_cliente_fornecedor_integracao": _get_row_value(row, header_map, "destinatario_cnpj"),
        "data_vencimento": _iso_to_br(_get_row_value(row, header_map, "vencimento")),
        "data_previsao": _iso_to_br(_get_row_value(row, header_map, "vencimento")),
        "valor_documento": _safe_float(_get_row_value(row, header_map, "valor_total_nf")),
        "codigo_categoria": codigo_categoria,
        "id_conta_corrente": int(id_conta_corrente),
        "numero_documento": _get_row_value(row, header_map, "numero_nf"),
        "numero_documento_fiscal": _get_row_value(row, header_map, "numero_nf"),
    }
    if payload["codigo_cliente_fornecedor"] is None:
        payload.pop("codigo_cliente_fornecedor")
    return _omie_request(OMIE_CONTARECEBER_URL, "IncluirContaReceber", payload, rate_limiter, prefix)


def process_xlsx_to_omie(xlsx_path: str):
    logger = get_logger("omie")
    _load_env()
    prefix = os.getenv("OMIE_PREFIX", "CUSTOMER").strip().upper()
    workbook = load_workbook(xlsx_path)
    if "contas_a_receber" not in workbook.sheetnames:
        raise ValueError("A aba 'contas_a_receber' não foi encontrada na planilha.")
    worksheet = workbook["contas_a_receber"]
    rows = list(worksheet.iter_rows(values_only=True))
    if not rows:
        raise ValueError("A aba 'contas_a_receber' está vazia.")
    headers = [str(h).strip() for h in rows[0]]
    header_map = {h: idx for idx, h in enumerate(headers)}
    rate_limiter = RateLimiter(MIN_INTERVAL_SECONDS)
    results = []
    data_rows = [row for row in rows[1:] if row and any(row)]
    log_section(logger, "Planilha carregada")
    log_status(logger, "ok", f"Arquivo: {xlsx_path}")
    log_status(logger, "ok", f"Linhas de dados: {len(data_rows)}")
    for row in data_rows:
        payload = _build_cliente_payload(row, header_map)
        cnpj_cpf = payload.get("cnpj_cpf")
        if not cnpj_cpf:
            log_status(logger, "warn", "Linha ignorada por ausência de CNPJ/CPF do destinatário.")
            continue
        cliente_resp = _upsert_cliente(payload, rate_limiter, prefix)
        codigo_cliente = cliente_resp.get("codigo_cliente_omie") or cliente_resp.get("codigo_cliente")
        conta_resp = _incluir_conta_receber(row, header_map, codigo_cliente, rate_limiter, prefix)
        results.append({"cliente": cliente_resp, "conta_receber": conta_resp})
    log_section(logger, "Envio concluído")
    log_status(logger, "ok", f"Registros processados: {len(results)}")
    return results
